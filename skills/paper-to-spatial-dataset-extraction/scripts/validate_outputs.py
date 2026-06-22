#!/usr/bin/env python3
"""Validate STRAND paper-to-dataset extraction run outputs."""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - reported as validation error.
    load_workbook = None


SKILL_DIR = Path(__file__).resolve().parents[1]
FIELDS_REFERENCE = SKILL_DIR / "references" / "strand_dataset_fields.md"
LEGACY_REQUIRED_FILES = [
    "dataset_rows.json",
    "dataset_rows.tsv",
    "Datasets_filled.xlsx",
    "unresolved_fields.tsv",
    "curation_report.md",
]
PUBLIC_REQUIRED_FILES = [
    "final_metadata.tsv",
    "final_metadata.xlsx",
    "extraction_report.md",
]
PUBLIC_FILES = set(PUBLIC_REQUIRED_FILES)
ALLOWED_PUBLIC_ROOT_FILES = PUBLIC_FILES | {"README.md"}
REPORT_SECTIONS = [
    "## Final Metadata Summary",
    "## Upstream Paper Search",
    "## Evidence Sources",
    "## Downloaded Files And Checksums",
    "## Field-Level Evidence",
    "## QC Parameters Used",
    "## QC And Count Convention Notes",
    "## Unresolved Or Curator-Review Fields",
    "## Review Conclusion",
    "## Internal Audit Files",
]
QC_PARAMETER_FIELDS = [
    "dataset",
    "sample",
    "step",
    "parameter",
    "value",
    "applied",
    "evidence_source",
    "effect_on_counts",
    "notes",
]


def target_fields():
    text = FIELDS_REFERENCE.read_text()
    match = re.search(r"```text\n(.*?)\n```", text, flags=re.S)
    if not match:
        raise SystemExit(f"could not parse target fields from {FIELDS_REFERENCE}")
    return [line.strip() for line in match.group(1).splitlines() if line.strip()]


def output_dir(path):
    path = Path(path)
    if path.name == "outputs":
        return path
    if (path / "final_metadata.tsv").exists() or (path / "final_metadata.xlsx").exists() or (path / "internal").exists():
        return path
    return path / "outputs"


def find_artifact(out_dir, name):
    for candidate in (out_dir / name, out_dir / "internal" / name):
        if candidate.exists():
            return candidate
    return out_dir / name


def read_tsv_header(path):
    with path.open(newline="") as handle:
        return next(csv.reader(handle, delimiter="\t"))


def count_unresolved(path):
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = list(reader)
    if rows and "field" not in (reader.fieldnames or []):
        raise ValueError("unresolved_fields.tsv must include a field column when unresolved rows exist")
    if rows and not ({"reason", "status"} & set(reader.fieldnames or [])):
        raise ValueError("unresolved_fields.tsv must include reason or status when unresolved rows exist")
    if rows and not ({"next_action", "needed_source"} & set(reader.fieldnames or [])):
        raise ValueError("unresolved_fields.tsv must include next_action or needed_source when unresolved rows exist")
    return len(rows)


def count_qc_parameters(path):
    with path.open(newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        rows = list(reader)
    if reader.fieldnames != QC_PARAMETER_FIELDS:
        raise ValueError(f"qc_parameters.tsv header must be {QC_PARAMETER_FIELDS}")
    if not rows:
        raise ValueError("qc_parameters.tsv must include at least one row")
    malformed_rows = [
        index
        for index, row in enumerate(rows, 1)
        if None in row or any(row.get(field) is None for field in QC_PARAMETER_FIELDS)
    ]
    if malformed_rows:
        raise ValueError(f"qc_parameters.tsv rows have the wrong number of columns: {malformed_rows}")
    empty_steps = [index for index, row in enumerate(rows, 1) if not row.get("step")]
    if empty_steps:
        raise ValueError(f"qc_parameters.tsv rows missing step: {empty_steps}")
    return len(rows)


def validate_upstream_search(path):
    problems = []
    try:
        data = json.loads(path.read_text())
    except Exception as exc:
        return [f"internal/upstream_paper_search.json is not valid JSON: {exc}"]

    if not isinstance(data, dict):
        return ["internal/upstream_paper_search.json must contain an object"]
    if "search_was_needed" not in data:
        problems.append("internal/upstream_paper_search.json missing search_was_needed")
    if data.get("search_was_needed"):
        queries = data.get("queries")
        if not isinstance(queries, list) or not queries:
            problems.append("search runs must record at least one query in upstream_paper_search.json")
        candidates = data.get("candidate_papers")
        if not isinstance(candidates, list):
            problems.append("search runs must record candidate_papers as a list")
        selected = data.get("selected_paper")
        if selected is not None and not isinstance(selected, dict):
            problems.append("selected_paper must be an object when present")
    return problems


def validate_xlsx_header(path, fields, label):
    problems = []
    if load_workbook is None:
        return [f"openpyxl is not importable; cannot open {label}"]
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        xlsx_header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if xlsx_header != fields:
            problems.append(f"{label} header does not match target field order")
        if "Catagory" in xlsx_header:
            problems.append(f"{label} uses legacy Catagory spelling")
    except Exception as exc:
        problems.append(f"could not open {label}: {exc}")
    return problems


def xlsx_nonblank_no_rows(path):
    if load_workbook is None:
        return []
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if "No." not in header:
        return []
    no_index = header.index("No.") + 1
    rows = []
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
        value = row[no_index - 1]
        if str(value or "").strip() not in {"", "-"}:
            rows.append(index)
    return rows


def validate_legacy(path, require_truth):
    out_dir = output_dir(path)
    fields = target_fields()
    problems = []

    for name in LEGACY_REQUIRED_FILES:
        if not find_artifact(out_dir, name).exists():
            problems.append(f"missing {name}")
    if require_truth and not find_artifact(out_dir, "truth_comparison.md").exists():
        problems.append("missing truth_comparison.md")
    if problems:
        return out_dir, problems, {}

    try:
        rows = json.loads(find_artifact(out_dir, "dataset_rows.json").read_text())
    except Exception as exc:
        problems.append(f"dataset_rows.json is not valid JSON: {exc}")
        rows = []
    legacy_wrapped_rows = False
    if isinstance(rows, dict) and isinstance(rows.get("rows"), list):
        legacy_wrapped_rows = True
        rows = rows["rows"]
    if not isinstance(rows, list):
        problems.append("dataset_rows.json must contain a list or legacy object with rows list")
        rows = []

    for index, row in enumerate(rows, 1):
        if not isinstance(row, dict):
            problems.append(f"row {index} is not an object")
            continue
        missing = [field for field in fields if field not in row]
        if missing:
            problems.append(f"row {index} missing fields: {missing}")
        if "Catagory" in row:
            problems.append(f"row {index} uses legacy Catagory spelling")

    try:
        tsv_header = read_tsv_header(find_artifact(out_dir, "dataset_rows.tsv"))
        if tsv_header != fields:
            problems.append("dataset_rows.tsv header does not match target field order")
        if "Catagory" in tsv_header:
            problems.append("dataset_rows.tsv uses legacy Catagory spelling")
    except Exception as exc:
        problems.append(f"could not read dataset_rows.tsv: {exc}")

    problems.extend(validate_xlsx_header(find_artifact(out_dir, "Datasets_filled.xlsx"), fields, "Datasets_filled.xlsx"))

    try:
        unresolved_count = count_unresolved(find_artifact(out_dir, "unresolved_fields.tsv"))
    except Exception as exc:
        unresolved_count = None
        problems.append(f"unresolved_fields.tsv is not auditable: {exc}")

    return out_dir, problems, {
        "mode": "legacy",
        "rows": len(rows),
        "unresolved_rows": unresolved_count,
        "truth_comparison": find_artifact(out_dir, "truth_comparison.md").exists(),
        "legacy_wrapped_rows": legacy_wrapped_rows,
    }


def validate_public(path, require_truth, external_user, require_upstream_search=False):
    out_dir = output_dir(path)
    fields = target_fields()
    problems = []
    internal_dir = out_dir / "internal"

    for name in PUBLIC_REQUIRED_FILES:
        if not (out_dir / name).exists():
            problems.append(f"missing {name}")
    if not internal_dir.exists():
        problems.append("missing internal directory")
    if problems:
        return out_dir, problems, {}

    try:
        tsv_header = read_tsv_header(out_dir / "final_metadata.tsv")
        if tsv_header != fields:
            problems.append("final_metadata.tsv header does not match target field order")
        if "Catagory" in tsv_header:
            problems.append("final_metadata.tsv uses legacy Catagory spelling")
    except Exception as exc:
        problems.append(f"could not read final_metadata.tsv: {exc}")

    problems.extend(validate_xlsx_header(out_dir / "final_metadata.xlsx", fields, "final_metadata.xlsx"))

    report_text = (out_dir / "extraction_report.md").read_text()
    for section in REPORT_SECTIONS:
        if section not in report_text:
            problems.append(f"extraction_report.md missing section: {section}")

    unresolved_path = internal_dir / "unresolved_fields.tsv"
    if unresolved_path.exists():
        try:
            unresolved_count = count_unresolved(unresolved_path)
        except Exception as exc:
            unresolved_count = None
            problems.append(f"internal/unresolved_fields.tsv is not auditable: {exc}")
    else:
        unresolved_count = None
        problems.append("missing internal/unresolved_fields.tsv")

    if not (internal_dir / "download_plan.md").exists():
        problems.append("missing internal/download_plan.md")

    qc_parameters_path = internal_dir / "qc_parameters.tsv"
    if qc_parameters_path.exists():
        try:
            qc_parameter_count = count_qc_parameters(qc_parameters_path)
        except Exception as exc:
            qc_parameter_count = None
            problems.append(f"internal/qc_parameters.tsv is not auditable: {exc}")
    else:
        qc_parameter_count = None
        problems.append("missing internal/qc_parameters.tsv")

    source_candidates = [
        internal_dir / "article_metadata.json",
        internal_dir / "source_metadata_inspection.json",
        internal_dir / "dataset_rows.json",
    ]
    if not any(path.exists() for path in source_candidates):
        problems.append("internal directory lacks source evidence JSON")

    if require_truth and not (internal_dir / "truth_comparison.md").exists():
        problems.append("missing internal/truth_comparison.md")

    upstream_search_path = internal_dir / "upstream_paper_search.json"
    if upstream_search_path.exists():
        problems.extend(validate_upstream_search(upstream_search_path))
    elif require_upstream_search:
        problems.append("missing internal/upstream_paper_search.json")

    row_count = 0
    try:
        with (out_dir / "final_metadata.tsv").open(newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter="\t"))
            row_count = len(rows)
        if external_user:
            numbered_rows = [
                index
                for index, row in enumerate(rows, 1)
                if str(row.get("No.", "")).strip() not in {"", "-"}
            ]
            if numbered_rows:
                problems.append(f"external-user output should leave No. blank unless user supplied IDs: rows {numbered_rows}")
            xlsx_numbered_rows = xlsx_nonblank_no_rows(out_dir / "final_metadata.xlsx")
            if xlsx_numbered_rows:
                problems.append(f"external-user XLSX should leave No. blank unless user supplied IDs: rows {xlsx_numbered_rows}")
    except Exception:
        pass

    root_audit_files = [
        path.name
        for path in out_dir.iterdir()
        if path.is_file() and path.name not in ALLOWED_PUBLIC_ROOT_FILES
    ]
    if root_audit_files:
        problems.append(f"audit files should be under internal, found at public root: {root_audit_files}")

    return out_dir, problems, {
        "mode": "public",
        "rows": row_count,
        "unresolved_rows": unresolved_count,
        "qc_parameter_rows": qc_parameter_count,
        "upstream_search": upstream_search_path.exists(),
        "truth_comparison": (internal_dir / "truth_comparison.md").exists(),
        "legacy_wrapped_rows": False,
    }


def validate(path, require_truth, mode, external_user, require_upstream_search=False):
    out_dir = output_dir(path)
    if mode == "legacy":
        return validate_legacy(path, require_truth)
    if mode == "public":
        return validate_public(path, require_truth, external_user, require_upstream_search)
    if (out_dir / "final_metadata.tsv").exists() or (out_dir / "final_metadata.xlsx").exists():
        return validate_public(path, require_truth, external_user, require_upstream_search)
    return validate_legacy(path, require_truth)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("paths", nargs="+", help="Run directories or outputs directories to validate")
    parser.add_argument("--require-truth", action="store_true", help="Require truth_comparison.md")
    parser.add_argument(
        "--external-user",
        action="store_true",
        help="Require external-user behavior, including blank No. values unless IDs were supplied",
    )
    parser.add_argument(
        "--mode",
        choices=["auto", "public", "legacy"],
        default="auto",
        help="Validate the public distributable package, legacy pilot outputs, or auto-detect",
    )
    parser.add_argument(
        "--require-upstream-search",
        action="store_true",
        help="Require internal/upstream_paper_search.json for search-first runs",
    )
    args = parser.parse_args()

    any_problems = False
    for path in args.paths:
        out_dir, problems, summary = validate(
            path,
            args.require_truth,
            args.mode,
            args.external_user,
            args.require_upstream_search,
        )
        if problems:
            any_problems = True
            print(f"FAIL {out_dir}")
            for problem in problems:
                print(f"  - {problem}")
        else:
            print(
                f"PASS {out_dir} "
                f"mode={summary['mode']} "
                f"rows={summary['rows']} "
                f"unresolved_rows={summary['unresolved_rows']} "
                f"truth_comparison={summary['truth_comparison']} "
                f"legacy_wrapped_rows={summary['legacy_wrapped_rows']}"
            )
    return 1 if any_problems else 0


if __name__ == "__main__":
    sys.exit(main())
